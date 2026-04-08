# -*- coding: utf-8 -*-
"""
发票数据动态仪表盘一键生成脚本
用法: python generate_invoice_dashboard.py <Excel文件路径>
示例: python generate_invoice_dashboard.py 发票清单.xlsx

功能:
1. 自动检测数据行数和列结构
2. 添加年份/月份辅助列
3. 创建动态仪表盘（SUMIF/SUMIFS公式）
4. 自动统计销方TOP10、税率分布、发票类型
"""

import sys
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from datetime import datetime

def analyze_data(ws):
    """分析Sheet1数据结构"""
    # 获取数据范围
    max_row = ws.max_row
    max_col = ws.max_column
    
    # 获取列标题
    headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]
    
    # 查找关键列索引
    col_map = {}
    for i, h in enumerate(headers, 1):
        if h and isinstance(h, str):
            col_map[h.strip()] = i
    
    return {
        'max_row': max_row,
        'max_col': max_col,
        'headers': headers,
        'col_map': col_map
    }

def get_unique_values(ws, col_idx, start_row, end_row):
    """获取某列的唯一值列表"""
    values = {}
    for r in range(start_row, end_row + 1):
        v = ws.cell(r, col_idx).value
        if v:
            values[str(v).strip()] = values.get(str(v).strip(), 0) + 1
    return list(values.keys())

def get_top_sellers(ws, col_idx, amt_col_idx, start_row, end_row, top_n=10):
    """获取销方TOP N"""
    seller_amts = {}
    for r in range(start_row, end_row + 1):
        seller = ws.cell(r, col_idx).value
        amt = ws.cell(r, amt_col_idx).value
        if seller and amt:
            seller = str(seller).strip()
            try:
                amt = float(amt)
            except:
                amt = 0
            seller_amts[seller] = seller_amts.get(seller, 0) + amt
    
    # 排序取TOP N
    sorted_sellers = sorted(seller_amts.items(), key=lambda x: x[1], reverse=True)[:top_n]
    return [s[0] for s in sorted_sellers]

def create_dashboard(wb, data_info):
    """创建动态仪表盘"""
    ws = wb.create_sheet('动态仪表盘')
    
    max_row = data_info['max_row']
    col_map = data_info['col_map']
    
    # 获取列索引
    date_col = col_map.get('开票日期', 1)
    invoice_col = col_map.get('发票号码', 2)
    buyer_col = col_map.get('购方名称', 3)
    seller_col = col_map.get('销方名称', 4)
    item_col = col_map.get('项目名称', 5)
    rate_col = col_map.get('税率', 6)
    amt_col = col_map.get('金额', 7)
    tax_col = col_map.get('税额', 8)
    total_col = col_map.get('价税合计', 9)
    type_col = col_map.get('发票类型', 10)
    
    # 数据范围
    data_range = f"$2:${max_row}"
    
    # 样式定义
    title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='微软雅黑', size=11, bold=True)
    kpi_font = Font(name='微软雅黑', size=14, bold=True, color='2F5496')
    normal_font = Font(name='微软雅黑', size=10)
    
    title_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_fill = PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid')
    kpi_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # ========== 1. 标题 ==========
    ws.merge_cells('A1:L1')
    ws['A1'] = '📊 发票数据动态仪表盘'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 35
    
    # ========== 2. KPI 指标卡 (行3-5) ==========
    kpi_items = [
        ('B', '💰 总价税合计', f'=SUM(Sheet1!${chr(64+total_col)}$2:${chr(64+total_col)}${max_row})'),
        ('D', '💵 发票总金额', f'=SUM(Sheet1!${chr(64+amt_col)}$2:${chr(64+amt_col)}${max_row})'),
        ('F', '🧾 税额合计', f'=SUM(Sheet1!${chr(64+tax_col)}$2:${chr(64+tax_col)}${max_row})'),
        ('H', '📋 发票张数', f'=COUNTA(Sheet1!${chr(64+invoice_col)}$2:${chr(64+invoice_col)}${max_row})'),
        ('J', '🏢 供应商数', f'=SUMPRODUCT(1/COUNTIF(Sheet1!${chr(64+seller_col)}$2:${chr(64+seller_col)}${max_row},Sheet1!${chr(64+seller_col)}$2:${chr(64+seller_col)}${max_row}))'),
        ('L', '📦 项目品类', f'=SUMPRODUCT(1/COUNTIF(Sheet1!${chr(64+item_col)}$2:${chr(64+item_col)}${max_row},Sheet1!${chr(64+item_col)}$2:${chr(64+item_col)}${max_row}))'),
    ]
    
    for col, label, formula in kpi_items:
        # 标签
        ws[f'{col}3'] = label
        ws[f'{col}3'].font = header_font
        ws[f'{col}3'].fill = header_fill
        ws[f'{col}3'].alignment = center_align
        ws[f'{col}3'].border = thin_border
        
        # 数值 - 使用较小字体确保显示
        ws[f'{col}5'] = formula
        ws[f'{col}5'].font = Font(name='微软雅黑', size=10, bold=True, color='2F5496')
        ws[f'{col}5'].fill = kpi_fill
        ws[f'{col}5'].alignment = center_align
        ws[f'{col}5'].border = thin_border
        # 使用紧凑格式，避免####显示
        if col in ['B', 'D', 'F']:
            ws[f'{col}5'].number_format = '#,##0'
        else:
            ws[f'{col}5'].number_format = '0'
    
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[5].height = 28
    
    # ========== 3. 月度趋势分析 (行7-17) ==========
    ws['B7'] = '📈 月度趋势分析'
    ws['B7'].font = Font(name='微软雅黑', size=12, bold=True, color='2F5496')
    
    # 表头
    month_headers = ['月份', '发票张数', '金额', '税额', '价税合计', '环比增幅']
    for i, h in enumerate(month_headers):
        cell = ws.cell(row=9, column=2+i)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # 获取所有年月
    ws1 = wb['Sheet1']
    year_month_data = {}
    for r in range(2, max_row + 1):
        date_val = ws1.cell(r, date_col).value
        if date_val:
            try:
                if isinstance(date_val, str):
                    dt = datetime.strptime(date_val.split()[0], '%Y-%m-%d')
                else:
                    dt = date_val
                ym = (dt.year, dt.month)
                year_month_data[ym] = True
            except:
                pass
    
    sorted_ym = sorted(year_month_data.keys())
    
    # 写入月度数据
    row_idx = 10
    for year, month in sorted_ym:
        # 月份标签
        ws.cell(row=row_idx, column=2).value = f'{year}-{month:02d}'
        ws.cell(row=row_idx, column=2).alignment = center_align
        ws.cell(row=row_idx, column=2).border = thin_border
        
        # 发票张数 - 直接用 YEAR/MONTH 从日期列提取
        ws.cell(row=row_idx, column=3).value = (
            f'=SUMPRODUCT((YEAR(Sheet1!$A$2:$A${max_row})={year})'
            f'*(MONTH(Sheet1!$A$2:$A${max_row})={month}))'
        )
        ws.cell(row=row_idx, column=3).alignment = center_align
        ws.cell(row=row_idx, column=3).border = thin_border
        
        # 金额
        ws.cell(row=row_idx, column=4).value = (
            f'=SUMPRODUCT((YEAR(Sheet1!$A$2:$A${max_row})={year})'
            f'*(MONTH(Sheet1!$A$2:$A${max_row})={month})'
            f'*Sheet1!${chr(64+amt_col)}$2:${chr(64+amt_col)}${max_row})'
        )
        ws.cell(row=row_idx, column=4).number_format = '#,##0'
        ws.cell(row=row_idx, column=4).alignment = center_align
        ws.cell(row=row_idx, column=4).border = thin_border
        
        # 税额
        ws.cell(row=row_idx, column=5).value = (
            f'=SUMPRODUCT((YEAR(Sheet1!$A$2:$A$115)={year})'
            f'*(MONTH(Sheet1!$A$2:$A$115)={month})'
            f'*Sheet1!${chr(64+tax_col)}$2:${chr(64+tax_col)}${max_row})'
        )
        ws.cell(row=row_idx, column=5).number_format = '#,##0'
        ws.cell(row=row_idx, column=5).alignment = center_align
        ws.cell(row=row_idx, column=5).border = thin_border
        
        # 价税合计
        ws.cell(row=row_idx, column=6).value = (
            f'=SUMPRODUCT((YEAR(Sheet1!$A$2:$A$115)={year})'
            f'*(MONTH(Sheet1!$A$2:$A$115)={month})'
            f'*Sheet1!${chr(64+total_col)}$2:${chr(64+total_col)}${max_row})'
        )
        ws.cell(row=row_idx, column=6).number_format = '#,##0'
        ws.cell(row=row_idx, column=6).alignment = center_align
        ws.cell(row=row_idx, column=6).border = thin_border
        
        # 环比增幅
        if row_idx > 10:
            ws.cell(row=row_idx, column=7).value = f'=IF(F{row_idx-1}=0,"-",F{row_idx}/F{row_idx-1}-1)'
        else:
            ws.cell(row=row_idx, column=7).value = '-'
        ws.cell(row=row_idx, column=7).number_format = '0.0%'
        ws.cell(row=row_idx, column=7).alignment = center_align
        ws.cell(row=row_idx, column=7).border = thin_border
        
        row_idx += 1
    
    # 合计行
    total_row = row_idx
    ws.cell(row=total_row, column=2).value = '合计'
    ws.cell(row=total_row, column=2).font = header_font
    ws.cell(row=total_row, column=2).fill = header_fill
    ws.cell(row=total_row, column=2).alignment = center_align
    ws.cell(row=total_row, column=2).border = thin_border
    
    for col in range(3, 7):
        ws.cell(row=total_row, column=col).value = f'=SUM({chr(64+col)}10:{chr(64+col)}{total_row-1})'
        ws.cell(row=total_row, column=col).font = header_font
        ws.cell(row=total_row, column=col).fill = header_fill
        ws.cell(row=total_row, column=col).alignment = center_align
        ws.cell(row=total_row, column=col).border = thin_border
        if col >= 4:
            ws.cell(row=total_row, column=col).number_format = '#,##0'
    
    ws.cell(row=total_row, column=7).value = '-'
    ws.cell(row=total_row, column=7).font = header_font
    ws.cell(row=total_row, column=7).fill = header_fill
    ws.cell(row=total_row, column=7).alignment = center_align
    ws.cell(row=total_row, column=7).border = thin_border
    
    # ========== 4. 购方分析 (行19起) ==========
    section_start = total_row + 3
    
    ws[f'B{section_start}'] = '🏢 购方分析'
    ws[f'B{section_start}'].font = Font(name='微软雅黑', size=12, bold=True, color='2F5496')
    
    buyers = get_unique_values(ws1, buyer_col, 2, max_row)
    
    # 表头
    for i, h in enumerate(['购方名称', '金额', '占比']):
        cell = ws.cell(row=section_start+1, column=2+i)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # 数据行
    for idx, buyer in enumerate(buyers):
        r = section_start + 2 + idx
        ws.cell(row=r, column=2).value = buyer
        ws.cell(row=r, column=2).alignment = left_align
        ws.cell(row=r, column=2).border = thin_border
        
        ws.cell(row=r, column=3).value = f'=SUMIF(Sheet1!${chr(64+buyer_col)}$2:${chr(64+buyer_col)}${max_row},B{r},Sheet1!${chr(64+total_col)}$2:${chr(64+total_col)}${max_row})'
        ws.cell(row=r, column=3).number_format = '#,##0'
        ws.cell(row=r, column=3).alignment = center_align
        ws.cell(row=r, column=3).border = thin_border
        
        ws.cell(row=r, column=4).value = f'=C{r}/$B$5'
        ws.cell(row=r, column=4).number_format = '0.0%'
        ws.cell(row=r, column=4).alignment = center_align
        ws.cell(row=r, column=4).border = thin_border
    
    # ========== 5. 销方TOP10 (右侧) ==========
    ws[f'G{section_start}'] = '🛒 销方 TOP10'
    ws[f'G{section_start}'].font = Font(name='微软雅黑', size=12, bold=True, color='2F5496')
    
    sellers = get_top_sellers(ws1, seller_col, total_col, 2, max_row, 10)
    
    # 表头
    for i, h in enumerate(['销方名称', '金额', '占比']):
        cell = ws.cell(row=section_start+1, column=7+i)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # 数据行
    for idx, seller in enumerate(sellers):
        r = section_start + 2 + idx
        ws.cell(row=r, column=7).value = seller
        ws.cell(row=r, column=7).alignment = left_align
        ws.cell(row=r, column=7).border = thin_border
        
        ws.cell(row=r, column=8).value = f'=SUMIF(Sheet1!${chr(64+seller_col)}$2:${chr(64+seller_col)}${max_row},G{r},Sheet1!${chr(64+total_col)}$2:${chr(64+total_col)}${max_row})'
        ws.cell(row=r, column=8).number_format = '#,##0'
        ws.cell(row=r, column=8).alignment = center_align
        ws.cell(row=r, column=8).border = thin_border
        
        ws.cell(row=r, column=9).value = f'=H{r}/$B$5'
        ws.cell(row=r, column=9).number_format = '0.0%'
        ws.cell(row=r, column=9).alignment = center_align
        ws.cell(row=r, column=9).border = thin_border
    
    # ========== 6. 税率分布 (下半部分) ==========
    # 计算正确的起始行：取购方和销方中较长的那个
    max_partner_rows = max(len(buyers), len(sellers))
    rate_section = section_start + max_partner_rows + 4
    
    ws[f'B{rate_section}'] = '📊 税率分布'
    ws[f'B{rate_section}'].font = Font(name='微软雅黑', size=12, bold=True, color='2F5496')
    
    rates = get_unique_values(ws1, rate_col, 2, max_row)
    
    # 表头
    for i, h in enumerate(['税率', '金额', '占比']):
        cell = ws.cell(row=rate_section+1, column=2+i)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # 数据行 - 使用 SUMPRODUCT+VALUE+LEFT 精确匹配百分比字符串
    rate_range = f'Sheet1!${chr(64+rate_col)}$2:${chr(64+rate_col)}${max_row}'
    total_range = f'Sheet1!${chr(64+total_col)}$2:${chr(64+total_col)}${max_row}'
    for idx, rate in enumerate(rates):
        r = rate_section + 2 + idx
        ws.cell(row=r, column=2).value = rate
        ws.cell(row=r, column=2).alignment = center_align
        ws.cell(row=r, column=2).border = thin_border
        
        # 提取百分比数值（去掉%符号转为数字）进行比较
        rate_num = rate.rstrip('%')
        ws.cell(row=r, column=3).value = (
            f'=SUMPRODUCT((VALUE(LEFT({rate_range},LEN({rate_range})-1))={rate_num})'
            f'*{total_range})'
        )
        ws.cell(row=r, column=3).number_format = '#,##0'
        ws.cell(row=r, column=3).alignment = center_align
        ws.cell(row=r, column=3).border = thin_border
        
        ws.cell(row=r, column=4).value = f'=C{r}/$B$5'
        ws.cell(row=r, column=4).number_format = '0.0%'
        ws.cell(row=r, column=4).alignment = center_align
        ws.cell(row=r, column=4).border = thin_border
    
    # ========== 7. 发票类型 (右侧) ==========
    ws[f'G{rate_section}'] = '📋 发票类型'
    ws[f'G{rate_section}'].font = Font(name='微软雅黑', size=12, bold=True, color='2F5496')
    
    types = get_unique_values(ws1, type_col, 2, max_row)
    
    # 表头
    for i, h in enumerate(['发票类型', '金额', '占比']):
        cell = ws.cell(row=rate_section+1, column=7+i)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # 数据行
    for idx, inv_type in enumerate(types):
        r = rate_section + 2 + idx
        ws.cell(row=r, column=7).value = inv_type
        ws.cell(row=r, column=7).alignment = center_align
        ws.cell(row=r, column=7).border = thin_border
        
        ws.cell(row=r, column=8).value = f'=SUMIF(Sheet1!${chr(64+type_col)}$2:${chr(64+type_col)}${max_row},G{r},Sheet1!${chr(64+total_col)}$2:${chr(64+total_col)}${max_row})'
        ws.cell(row=r, column=8).number_format = '#,##0'
        ws.cell(row=r, column=8).alignment = center_align
        ws.cell(row=r, column=8).border = thin_border
        
        ws.cell(row=r, column=9).value = f'=H{r}/$B$5'
        ws.cell(row=r, column=9).number_format = '0.0%'
        ws.cell(row=r, column=9).alignment = center_align
        ws.cell(row=r, column=9).border = thin_border
    
    # ========== 设置列宽和字体 ==========
    # 调整列宽，确保内容完整显示
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 32  # 加宽名称列
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 32  # 加宽名称列
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 3
    ws.column_dimensions['L'].width = 14
    
    # 设置默认行高，确保字体显示
    for row in range(1, 50):
        ws.row_dimensions[row].height = 18

def add_helper_columns(wb, data_info):
    """添加年份/月份辅助列到Sheet1，使用 YEAR()/MONTH() 公式"""
    ws = wb['Sheet1']
    max_row = data_info['max_row']
    col_map = data_info['col_map']
    date_col = col_map.get('开票日期', 1)
    date_col_letter = chr(64 + date_col)  # 列字母
    
    # 设置表头
    ws.cell(1, 11).value = '年份'
    ws.cell(1, 11).font = Font(bold=True)
    ws.cell(1, 12).value = '月份'
    ws.cell(1, 12).font = Font(bold=True)
    
    # 使用 YEAR()/MONTH() 公式从开票日期自动提取
    for r in range(2, max_row + 1):
        ws.cell(r, 11).value = f'=YEAR({date_col_letter}{r})'
        ws.cell(r, 12).value = f'=MONTH({date_col_letter}{r})'

def convert_numeric_columns(wb, data_info):
    """确保金额/税额/价税合计为数值类型"""
    ws = wb['Sheet1']
    max_row = data_info['max_row']
    col_map = data_info['col_map']
    
    for col_name in ['金额', '税额', '价税合计']:
        col_idx = col_map.get(col_name)
        if col_idx:
            for r in range(2, max_row + 1):
                val = ws.cell(r, col_idx).value
                if isinstance(val, str):
                    try:
                        ws.cell(r, col_idx).value = float(val.strip())
                    except:
                        pass

def main():
    if len(sys.argv) < 2:
        print('用法: python generate_invoice_dashboard.py <Excel文件路径>')
        print('示例: python generate_invoice_dashboard.py 发票清单.xlsx')
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    if not os.path.exists(file_path):
        print(f'错误: 文件不存在 - {file_path}')
        sys.exit(1)
    
    print(f'正在处理: {file_path}')
    
    # 加载文件
    wb = load_workbook(file_path)
    
    # 检查Sheet1
    if 'Sheet1' not in wb.sheetnames:
        print('错误: 未找到 Sheet1 工作表')
        sys.exit(1)
    
    # 分析数据
    ws1 = wb['Sheet1']
    data_info = analyze_data(ws1)
    print(f'数据行数: {data_info["max_row"]}')
    print(f'列: {data_info["headers"]}')
    
    # 转换数值列
    print('转换数值列...')
    convert_numeric_columns(wb, data_info)
    
    # 添加辅助列
    print('添加年月辅助列...')
    add_helper_columns(wb, data_info)
    
    # 删除旧仪表盘
    if '动态仪表盘' in wb.sheetnames:
        del wb['动态仪表盘']
        print('删除旧仪表盘...')
    
    # 创建仪表盘
    print('创建动态仪表盘...')
    create_dashboard(wb, data_info)
    
    # 保存
    wb.save(file_path)
    print(f'完成! 文件已保存: {file_path}')
    print('提示: 打开Excel后按 Ctrl+Alt+F9 刷新所有公式')

if __name__ == '__main__':
    main()
