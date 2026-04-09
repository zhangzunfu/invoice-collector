#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
发票归集软件主界面
"""

import sys
import os
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem, 
    QHeaderView, QTextEdit, QTabWidget, QProgressBar, QMessageBox,
    QLineEdit, QCheckBox, QGroupBox, QFormLayout, QComboBox, QFrame
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt6.QtGui import QFont, QIcon
import pandas as pd
from datetime import datetime, timedelta
import threading
import time
from openpyxl.styles import Font as ExcelFont, Alignment, PatternFill, Border, Side

from data_processing.invoice_parser import InvoiceProcessor


class InvoiceScannerThread(QThread):
    """发票扫描线程"""
    progress_update = pyqtSignal(int)
    status_update = pyqtSignal(str)
    scan_complete = pyqtSignal(list)
    log_update = pyqtSignal(str)
    
    def __init__(self, folder_path, auto_monitor=False):
        super().__init__()
        self.folder_path = folder_path
        self.auto_monitor = auto_monitor
        
    def run(self):
        try:
            processor = InvoiceProcessor()
            invoices = processor.process_folder(self.folder_path)
            
            # 按开票日期排序
            invoices.sort(key=lambda x: x.get('开票日期', ''), reverse=False)
            
            self.scan_complete.emit(invoices)
            self.status_update.emit(f"扫描完成，找到 {len(invoices)} 张发票")
            self.log_update.emit(f"扫描完成，共处理 {len(invoices)} 张发票")
        except Exception as e:
            self.status_update.emit(f"扫描失败: {str(e)}")
            self.log_update.emit(f"扫描失败: {str(e)}")


class DashboardWidget(QWidget):
    """智能仪表盘组件 - 增强版"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_invoices = []
        self.df = None
        self.init_ui()
        self.init_charts()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        # 标题
        title_label = QLabel("发票智能统计仪表盘")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("""
            font-size: 20px; 
            font-weight: bold; 
            margin: 10px;
            color: #2c3e50;
        """)
        
        # 时间筛选区域
        filter_group = QGroupBox("时间筛选")
        filter_layout = QHBoxLayout()
        
        self.start_date_label = QLabel("开始:")
        # 年份选择
        self.start_year_combo = QComboBox()
        self.start_year_combo.setMinimumWidth(80)
        self.start_year_combo.setStyleSheet("padding: 2px;")
        # 月份选择
        self.start_month_combo = QComboBox()
        self.start_month_combo.setMinimumWidth(60)
        self.start_month_combo.setStyleSheet("padding: 2px;")
        
        self.end_date_label = QLabel("结束:")
        # 年份选择
        self.end_year_combo = QComboBox()
        self.end_year_combo.setMinimumWidth(80)
        self.end_year_combo.setStyleSheet("padding: 2px;")
        # 月份选择
        self.end_month_combo = QComboBox()
        self.end_month_combo.setMinimumWidth(60)
        self.end_month_combo.setStyleSheet("padding: 2px;")
        
        self.apply_filter_btn = QPushButton("应用筛选")
        self.apply_filter_btn.clicked.connect(self.apply_filter)
        self.apply_filter_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        
        self.clear_filter_btn = QPushButton("清除筛选")
        self.clear_filter_btn.clicked.connect(self.clear_filter)
        self.clear_filter_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        
        filter_layout.addWidget(self.start_date_label)
        filter_layout.addWidget(self.start_year_combo)
        filter_layout.addWidget(QLabel("年"))
        filter_layout.addWidget(self.start_month_combo)
        filter_layout.addWidget(QLabel("月"))
        filter_layout.addWidget(self.end_date_label)
        filter_layout.addWidget(self.end_year_combo)
        filter_layout.addWidget(QLabel("年"))
        filter_layout.addWidget(self.end_month_combo)
        filter_layout.addWidget(QLabel("月"))
        filter_layout.addWidget(self.apply_filter_btn)
        filter_layout.addWidget(self.clear_filter_btn)
        filter_layout.addStretch()
        filter_group.setLayout(filter_layout)
        
        # 统计卡片区域
        stats_group = QGroupBox("概览统计")
        stats_layout = QHBoxLayout()
        
        # 总发票数
        self.total_count_card = self.create_stat_card("总发票数", "0", "#3498db")
        stats_layout.addWidget(self.total_count_card)
        
        # 总金额
        self.total_amount_card = self.create_stat_card("总金额", "¥0", "#e74c3c")
        stats_layout.addWidget(self.total_amount_card)
        
        # 总税额
        self.total_tax_card = self.create_stat_card("总税额", "¥0", "#e67e22")
        stats_layout.addWidget(self.total_tax_card)
        
        # 本月发票数
        self.monthly_count_card = self.create_stat_card("本月发票", "0", "#27ae60")
        stats_layout.addWidget(self.monthly_count_card)
        
        # 本月金额
        self.monthly_amount_card = self.create_stat_card("本月金额", "¥0", "#f39c12")
        stats_layout.addWidget(self.monthly_amount_card)
        
        # 本月税额
        self.monthly_tax_card = self.create_stat_card("本月税额", "¥0", "#d35400")
        stats_layout.addWidget(self.monthly_tax_card)
        
        # 销方数量
        self.suppliers_card = self.create_stat_card("销方数量", "0", "#9b59b6")
        stats_layout.addWidget(self.suppliers_card)
        
        # 购方数量
        self.buyers_card = self.create_stat_card("购方数量", "0", "#1abc9c")
        stats_layout.addWidget(self.buyers_card)
        
        stats_layout.addStretch()
        stats_group.setLayout(stats_layout)
        
        # 图表区域
        charts_group = QGroupBox("详细分析")
        charts_layout = QVBoxLayout()
        
        # 月度趋势图 - 扩大显示区域
        self.monthly_trend_frame = QFrame()
        self.monthly_trend_layout = QVBoxLayout(self.monthly_trend_frame)
        self.monthly_trend_layout.setContentsMargins(0, 0, 0, 0)
        self.monthly_trend_label = QLabel("月度金额趋势")
        self.monthly_trend_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50; margin-bottom: 5px;")
        self.monthly_trend_chart_placeholder = QLabel("等待数据...")
        self.monthly_trend_chart_placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.monthly_trend_chart_placeholder.setStyleSheet("background-color: white; border: 1px solid #ddd; padding: 20px; color: #7f8c8d;")
        self.monthly_trend_layout.addWidget(self.monthly_trend_label)
        self.monthly_trend_layout.addWidget(self.monthly_trend_chart_placeholder)
        self.monthly_trend_layout.setStretch(1, 1)  # 让图表占满剩余空间
        
        # 销方排行榜和发票类型分布的水平布局
        charts_horizontal = QHBoxLayout()
        
        # 销方排行榜
        self.supplier_ranking_frame = QFrame()
        self.supplier_ranking_layout = QVBoxLayout(self.supplier_ranking_frame)
        self.supplier_ranking_layout.setContentsMargins(0, 0, 0, 0)
        self.supplier_ranking_label = QLabel("销方排行榜（按金额）")
        self.supplier_ranking_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50; margin-bottom: 5px;")
        self.supplier_ranking_chart_placeholder = QLabel("等待数据...")
        self.supplier_ranking_chart_placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.supplier_ranking_chart_placeholder.setStyleSheet("background-color: white; border: 1px solid #ddd; padding: 20px; color: #7f8c8d;")
        self.supplier_ranking_layout.addWidget(self.supplier_ranking_label)
        self.supplier_ranking_layout.addWidget(self.supplier_ranking_chart_placeholder)
        self.supplier_ranking_layout.setStretch(1, 1)
        
        # 发票类型分布
        self.type_distribution_frame = QFrame()
        self.type_distribution_layout = QVBoxLayout(self.type_distribution_frame)
        self.type_distribution_layout.setContentsMargins(0, 0, 0, 0)
        self.type_distribution_label = QLabel("发票类型分布")
        self.type_distribution_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50; margin-bottom: 5px;")
        self.type_distribution_chart_placeholder = QLabel("等待数据...")
        self.type_distribution_chart_placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.type_distribution_chart_placeholder.setStyleSheet("background-color: white; border: 1px solid #ddd; padding: 20px; color: #7f8c8d;")
        self.type_distribution_layout.addWidget(self.type_distribution_label)
        self.type_distribution_layout.addWidget(self.type_distribution_chart_placeholder)
        self.type_distribution_layout.setStretch(1, 1)
        
        charts_horizontal.addWidget(self.supplier_ranking_frame)
        charts_horizontal.addWidget(self.type_distribution_frame)
        charts_horizontal.setStretch(0, 1)
        charts_horizontal.setStretch(1, 1)
        
        charts_layout.addWidget(self.monthly_trend_frame)
        charts_layout.addLayout(charts_horizontal)
        charts_layout.setStretch(0, 3)  # 趋势图占3份
        charts_layout.setStretch(1, 2)  # 其他图表占2份
        charts_group.setLayout(charts_layout)
        
        # 更新按钮
        refresh_btn = QPushButton("刷新数据")
        refresh_btn.clicked.connect(self.refresh_data)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        
        layout.addWidget(title_label)
        layout.addWidget(filter_group)
        layout.addWidget(stats_group)
        layout.addWidget(charts_group)
        layout.addWidget(refresh_btn)
        layout.addStretch()
        
        self.setLayout(layout)
    
    def create_stat_card(self, title, value, color):
        """创建统计卡片 - 紧凑版"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: white;
                border-radius: 4px;
                border: 1px solid #ecf0f1;
                padding: 8px;
            }}
        """)
        
        layout = QVBoxLayout()
        layout.setSpacing(2)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # 标题
        title_label = QLabel(title)
        title_label.setStyleSheet("""
            font-size: 11px;
            color: #95a5a6;
            font-weight: normal;
        """)
        
        # 数值
        value_label = QLabel(value)
        value_label.setStyleSheet(f"""
            font-size: 20px;
            color: {color};
            font-weight: bold;
        """)
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        layout.addWidget(title_label)
        layout.addWidget(value_label)
        layout.addStretch()
        
        card.setLayout(layout)
        
        # 保存引用以便更新
        card.value_label = value_label
        
        return card
    
    def init_charts(self):
        """初始化图表"""
        try:
            import matplotlib
            matplotlib.use('QtAgg')
            from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
            from matplotlib.figure import Figure
            import matplotlib.pyplot as plt
            
            # 设置中文字体
            plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'Arial Unicode MS']
            plt.rcParams['axes.unicode_minus'] = False
            
            self.matplotlib_available = True
            self.FigureCanvas = FigureCanvas
            self.Figure = Figure
            self.plt = plt
        except ImportError:
            self.matplotlib_available = False
    
    def update_stats(self, invoices, update_filter_combos=False):
        """更新统计信息
        
        Args:
            invoices: 发票数据列表
            update_filter_combos: 是否更新时间筛选下拉列表（默认为False，避免筛选后覆盖用户选择）
        """
        self.current_invoices = invoices
        
        if not invoices:
            # 清空所有统计数据
            self.total_count_card.value_label.setText("0")
            self.total_amount_card.value_label.setText("¥0")
            self.total_count_card.value_label.setText("0")
            self.monthly_amount_card.value_label.setText("¥0")
            self.suppliers_card.value_label.setText("0")
            self.buyers_card.value_label.setText("0")
            
            # 清空时间筛选
            if hasattr(self, 'start_year_combo'):
                self.start_year_combo.clear()
                self.start_month_combo.clear()
            if hasattr(self, 'end_year_combo'):
                self.end_year_combo.clear()
                self.end_month_combo.clear()
            
            # 清空图表
            self.clear_charts()
            return
        
        # 创建DataFrame用于分析
        self.df = pd.DataFrame(invoices)
        
        # 数据清洗
        self._clean_data()
        
        # 更新统计卡片
        self._update_summary_stats()
        
        # 更新图表
        self._update_charts()
        
        # 只在指定时更新时间筛选下拉列表
        if update_filter_combos:
            self._update_filter_combos(invoices)
            
            # 设置时间筛选默认值（第一张和最后一张发票的月份）
            if hasattr(self, 'start_year_combo') and hasattr(self, 'end_year_combo'):
                if len(invoices) > 0:
                    # 第一张发票的日期
                    first_date_str = invoices[0].get('开票日期', '')
                    # 最后一张发票的日期
                    last_date_str = invoices[-1].get('开票日期', '')
                    
                    # 转换日期
                    if first_date_str:
                        if '年' in first_date_str:
                            first_date = datetime.strptime(first_date_str, '%Y年%m月%d日')
                        else:
                            first_date = datetime.strptime(first_date_str, '%Y-%m-%d')
                        
                        # 设置开始年份和月份
                        start_year_idx = self.start_year_combo.findText(str(first_date.year))
                        if start_year_idx >= 0:
                            self.start_year_combo.setCurrentIndex(start_year_idx)
                        
                        start_month_idx = self.start_month_combo.findText(str(first_date.month))
                        if start_month_idx >= 0:
                            self.start_month_combo.setCurrentIndex(start_month_idx)
                    
                    if last_date_str:
                        if '年' in last_date_str:
                            last_date = datetime.strptime(last_date_str, '%Y年%m月%d日')
                        else:
                            last_date = datetime.strptime(last_date_str, '%Y-%m-%d')
                        
                        # 设置结束年份和月份
                        end_year_idx = self.end_year_combo.findText(str(last_date.year))
                        if end_year_idx >= 0:
                            self.end_year_combo.setCurrentIndex(end_year_idx)
                        
                        end_month_idx = self.end_month_combo.findText(str(last_date.month))
                        if end_month_idx >= 0:
                            self.end_month_combo.setCurrentIndex(end_month_idx)
    
    def _update_filter_combos(self, invoices):
        """更新时间筛选下拉列表"""
        if not invoices:
            return
        
        # 收集所有年份
        years = set()
        months = set()
        
        for invoice in invoices:
            date_str = invoice.get('开票日期', '')
            if not date_str:
                continue
            
            try:
                if '年' in date_str:
                    date = datetime.strptime(date_str, '%Y年%m月%d日')
                else:
                    date = datetime.strptime(date_str, '%Y-%m-%d')
                
                years.add(date.year)
                months.add(date.month)
            except:
                continue
        
        # 更新年份下拉列表
        if hasattr(self, 'start_year_combo'):
            current_start_year = self.start_year_combo.currentText()
            self.start_year_combo.clear()
            self.start_year_combo.addItem("全部")
            for year in sorted(years):
                self.start_year_combo.addItem(str(year))
            
            # 恢复之前的选择
            if current_start_year:
                idx = self.start_year_combo.findText(current_start_year)
                if idx >= 0:
                    self.start_year_combo.setCurrentIndex(idx)
        
        if hasattr(self, 'end_year_combo'):
            current_end_year = self.end_year_combo.currentText()
            self.end_year_combo.clear()
            self.end_year_combo.addItem("全部")
            for year in sorted(years):
                self.end_year_combo.addItem(str(year))
            
            # 恢复之前的选择
            if current_end_year:
                idx = self.end_year_combo.findText(current_end_year)
                if idx >= 0:
                    self.end_year_combo.setCurrentIndex(idx)
        
        # 更新月份下拉列表（1-12月）
        if hasattr(self, 'start_month_combo'):
            current_start_month = self.start_month_combo.currentText()
            self.start_month_combo.clear()
            self.start_month_combo.addItem("全部")
            for month in range(1, 13):
                self.start_month_combo.addItem(str(month))
            
            # 恢复之前的选择
            if current_start_month:
                idx = self.start_month_combo.findText(current_start_month)
                if idx >= 0:
                    self.start_month_combo.setCurrentIndex(idx)
        
        if hasattr(self, 'end_month_combo'):
            current_end_month = self.end_month_combo.currentText()
            self.end_month_combo.clear()
            self.end_month_combo.addItem("全部")
            for month in range(1, 13):
                self.end_month_combo.addItem(str(month))
            
            # 恢复之前的选择
            if current_end_month:
                idx = self.end_month_combo.findText(current_end_month)
                if idx >= 0:
                    self.end_month_combo.setCurrentIndex(idx)
    
    def _clean_data(self):
        """清洗数据"""
        if self.df is None or len(self.df) == 0:
            return
        
        # 转换日期格式 - 支持多种格式
        try:
            # 先尝试"YYYY-MM-DD"格式
            self.df['开票日期_dt'] = pd.to_datetime(self.df['开票日期'], format='%Y-%m-%d', errors='coerce')
            # 如果有失败，尝试"YYYY年MM月DD日"格式
            if self.df['开票日期_dt'].isna().any():
                mask = self.df['开票日期_dt'].isna()
                if mask.any():
                    self.df.loc[mask, '开票日期_dt'] = pd.to_datetime(
                        self.df.loc[mask, '开票日期'], format='%Y年%m月%d日', errors='coerce'
                    )
        except Exception as e:
            # 如果格式转换失败，使用自动检测
            try:
                self.df['开票日期_dt'] = pd.to_datetime(self.df['开票日期'], errors='coerce')
            except:
                pass
        
        # 转换金额字段为数值
        for col in ['金额', '税额', '价税合计']:
            if col in self.df.columns:
                self.df[col] = pd.to_numeric(self.df[col], errors='coerce').fillna(0)
        
        # 转换税率为数值（去除百分号）
        if '税率' in self.df.columns:
            self.df['税率_数值'] = self.df['税率'].astype(str).str.rstrip('%')
            self.df['税率_数值'] = pd.to_numeric(self.df['税率_数值'], errors='coerce').fillna(0)
        
        # 提取月份
        if '开票日期_dt' in self.df.columns:
            self.df['月份'] = self.df['开票日期_dt'].dt.strftime('%Y-%m')
    
    def _update_summary_stats(self):
        """更新概览统计"""
        if self.df is None or len(self.df) == 0:
            return
        
        # 总发票数
        total_count = len(self.df)
        self.total_count_card.value_label.setText(str(total_count))
        
        # 总金额
        total_amount = self.df['价税合计'].sum()
        self.total_amount_card.value_label.setText(f"¥{total_amount:,.2f}")
        
        # 总税额
        if '税额' in self.df.columns:
            total_tax = self.df['税额'].sum()
            self.total_tax_card.value_label.setText(f"¥{total_tax:,.2f}")
        else:
            self.total_tax_card.value_label.setText("¥0")
        
        # 本月发票数
        current_month = datetime.now().strftime('%Y-%m')
        if '月份' in self.df.columns:
            monthly_count = len(self.df[self.df['月份'] == current_month])
            self.monthly_count_card.value_label.setText(str(monthly_count))
        else:
            self.monthly_count_card.value_label.setText("0")
        
        # 本月金额
        if '月份' in self.df.columns:
            monthly_amount = self.df[self.df['月份'] == current_month]['价税合计'].sum()
            self.monthly_amount_card.value_label.setText(f"¥{monthly_amount:,.2f}")
        else:
            self.monthly_amount_card.value_label.setText("¥0")
        
        # 本月税额
        if '月份' in self.df.columns and '税额' in self.df.columns:
            monthly_tax = self.df[self.df['月份'] == current_month]['税额'].sum()
            self.monthly_tax_card.value_label.setText(f"¥{monthly_tax:,.2f}")
        else:
            self.monthly_tax_card.value_label.setText("¥0")
        
        # 销方数量
        suppliers = len(self.df['销方名称'].unique())
        self.suppliers_card.value_label.setText(str(suppliers))
        
        # 购方数量
        buyers = len(self.df['购方名称'].unique())
        self.buyers_card.value_label.setText(str(buyers))
    
    def _update_charts(self):
        """更新图表"""
        if not self.matplotlib_available:
            # matplotlib不可用，显示文本统计
            self._update_charts_as_text()
            return
        
        if self.df is None or len(self.df) == 0:
            self.clear_charts()
            return
        
        # 月度趋势图
        self._update_monthly_trend_chart()
        
        # 销方排行榜
        self._update_supplier_ranking_chart()
        
        # 发票类型分布
        self._update_type_distribution_chart()
    
    def _update_monthly_trend_chart(self):
        """更新月度趋势图 - 固定12个月，使用曲线，去掉上边框和左边框"""
        if not self.matplotlib_available or self.df is None:
            return
        
        if '月份' not in self.df.columns:
            self.monthly_trend_chart_placeholder.setText("数据不足")
            return
        
        try:
            # 清除旧的图表
            for i in reversed(range(self.monthly_trend_layout.count())):
                item = self.monthly_trend_layout.itemAt(i)
                if item.widget():
                    if item.widget() != self.monthly_trend_label:
                        item.widget().deleteLater()
            
            # 创建新图表
            figure = self.Figure(figsize=(10, 4), dpi=100)
            canvas = self.FigureCanvas(figure)
            ax = figure.add_subplot(111)
            
            # 生成过去12个月的月份列表
            from dateutil.relativedelta import relativedelta
            current_date = datetime.now()
            months_list = []
            for i in range(11, -1, -1):
                month_date = current_date - relativedelta(months=i)
                months_list.append(month_date.strftime('%Y-%m'))
            
            # 按月份汇总数据
            monthly = self.df.groupby('月份').agg({
                '价税合计': 'sum'
            }).reset_index()
            
            # 为12个月填充数据（没有数据的月份显示0）
            monthly_data = []
            for month in months_list:
                month_row = monthly[monthly['月份'] == month]
                if len(month_row) > 0:
                    monthly_data.append(float(month_row['价税合计'].values[0]))
                else:
                    monthly_data.append(0.0)
            
            # 绘制折线图
            x_data = range(len(months_list))  # 0-11
            y_data = monthly_data
            
            # 使用样条曲线绘制平滑曲线
            from scipy.interpolate import make_interp_spline
            import numpy as np
            
            # 创建平滑的曲线
            x_smooth = np.linspace(0, len(x_data) - 1, 200)
            spl = make_interp_spline(x_data, y_data, k=3)  # 三次样条插值
            y_smooth = spl(x_smooth)
            
            # 绘制平滑曲线
            ax.plot(x_smooth, y_smooth, linewidth=2, color='#3498db', alpha=0.8)
            
            # 绘制数据点
            ax.scatter(x_data, y_data, s=60, color='#3498db', zorder=5, alpha=0.9)
            
            # 设置x轴标签（固定显示1-12月）
            ax.set_xticks(x_data)
            ax.set_xticklabels([m.split('-')[1] + '月' for m in months_list], 
                              rotation=0, ha='center', fontsize=10)
            
            # 去掉上边框和左边框
            ax.spines['top'].set_visible(False)
            ax.spines['left'].set_visible(False)
            
            # 只保留底部和右边框
            ax.spines['bottom'].set_color('#ecf0f1')
            ax.spines['right'].set_color('#ecf0f1')
            ax.spines['bottom'].set_linewidth(0.5)
            ax.spines['right'].set_linewidth(0.5)
            
            # 隐藏y轴刻度
            ax.set_yticks([])
            
            # 在点上显示数值
            for i, (x, y) in enumerate(zip(x_data, y_data)):
                if y > 0:  # 只显示有数据的点
                    ax.text(x, y, f'{y:,.0f}', ha='center', va='bottom', fontsize=8, 
                           bbox=dict(boxstyle='round,pad=0.3', facecolor='white', 
                                   edgecolor='#3498db', alpha=0.8, linewidth=0.5))
            
            # 设置标题
            ax.set_title('月度金额趋势（过去12个月）', fontsize=12, fontweight='bold', 
                         pad=10, color='#2c3e50')
            
            # 去掉x轴和y轴标签
            ax.set_xlabel('')
            ax.set_ylabel('')
            
            # 去掉网格线
            ax.grid(False)
            
            figure.tight_layout()
            canvas.draw()
            
            self.monthly_trend_layout.addWidget(canvas)
            
        except ImportError:
            # 如果scipy不可用，使用普通曲线
            try:
                # 清除旧的图表
                for i in reversed(range(self.monthly_trend_layout.count())):
                    item = self.monthly_trend_layout.itemAt(i)
                    if item.widget():
                        if item.widget() != self.monthly_trend_label:
                            item.widget().deleteLater()
                
                # 创建新图表
                figure = self.Figure(figsize=(10, 4), dpi=100)
                canvas = self.FigureCanvas(figure)
                ax = figure.add_subplot(111)
                
                # 生成过去12个月的月份列表
                from dateutil.relativedelta import relativedelta
                current_date = datetime.now()
                months_list = []
                for i in range(11, -1, -1):
                    month_date = current_date - relativedelta(months=i)
                    months_list.append(month_date.strftime('%Y-%m'))
                
                # 按月份汇总数据
                monthly = self.df.groupby('月份').agg({
                    '价税合计': 'sum'
                }).reset_index()
                
                # 为12个月填充数据（没有数据的月份显示0）
                monthly_data = []
                for month in months_list:
                    month_row = monthly[monthly['月份'] == month]
                    if len(month_row) > 0:
                        monthly_data.append(float(month_row['价税合计'].values[0]))
                    else:
                        monthly_data.append(0.0)
                
                # 绘制折线图
                x_data = range(len(months_list))  # 0-11
                y_data = monthly_data
                
                # 绘制曲线
                ax.plot(x_data, y_data, marker='o', linewidth=2, markersize=6, 
                       color='#3498db', linestyle='-', markerfacecolor='#2980b9', markeredgewidth=2)
                
                # 设置x轴标签
                ax.set_xticks(x_data)
                ax.set_xticklabels([m.split('-')[1] + '月' for m in months_list], 
                                  rotation=0, ha='center', fontsize=10)
                
                # 去掉上边框和左边框
                ax.spines['top'].set_visible(False)
                ax.spines['left'].set_visible(False)
                
                # 只保留底部和右边框
                ax.spines['bottom'].set_color('#ecf0f1')
                ax.spines['right'].set_color('#ecf0f1')
                ax.spines['bottom'].set_linewidth(0.5)
                ax.spines['right'].set_linewidth(0.5)
                
                # 隐藏y轴刻度
                ax.set_yticks([])
                
                # 在点上显示数值
                for i, (x, y) in enumerate(zip(x_data, y_data)):
                    if y > 0:  # 只显示有数据的点
                        ax.text(x, y, f'{y:,.0f}', ha='center', va='bottom', fontsize=8,
                               bbox=dict(boxstyle='round,pad=0.3', facecolor='white', 
                                       edgecolor='#3498db', alpha=0.8, linewidth=0.5))
                
                # 设置标题
                ax.set_title('月度金额趋势（过去12个月）', fontsize=12, fontweight='bold', 
                             pad=10, color='#2c3e50')
                
                # 去掉x轴和y轴标签
                ax.set_xlabel('')
                ax.set_ylabel('')
                
                # 去掉网格线
                ax.grid(False)
                
                figure.tight_layout()
                canvas.draw()
                
                self.monthly_trend_layout.addWidget(canvas)
                
            except Exception as e:
                self.monthly_trend_chart_placeholder.setText(f"图表加载失败: {str(e)}")
        except Exception as e:
            self.monthly_trend_chart_placeholder.setText(f"图表加载失败: {str(e)}")
    
    def _update_supplier_ranking_chart(self):
        """更新销方排行榜"""
        if not self.matplotlib_available or self.df is None:
            return
        
        # 按销方汇总
        supplier_stats = self.df.groupby('销方名称').agg({
            '价税合计': 'sum'
        }).reset_index()
        supplier_stats = supplier_stats.sort_values('价税合计', ascending=False).head(10)
        
        if len(supplier_stats) == 0:
            self.supplier_ranking_chart_placeholder.setText("无数据")
            return
        
        # 创建图表
        try:
            # 清除旧的图表
            for i in reversed(range(self.supplier_ranking_layout.count())):
                item = self.supplier_ranking_layout.itemAt(i)
                if item.widget():
                    if item.widget() != self.supplier_ranking_label:
                        item.widget().deleteLater()
            
            # 创建新图表
            figure = self.Figure(figsize=(6, 4), dpi=100)
            canvas = self.FigureCanvas(figure)
            ax = figure.add_subplot(111)
            
            # 绘制水平条形图
            x_data = supplier_stats['价税合计'].tolist()
            y_data = supplier_stats['销方名称'].tolist()
            
            bars = ax.barh(range(len(y_data)), x_data, color='#e74c3c', alpha=0.7)
            ax.set_yticks(range(len(y_data)))
            ax.set_yticklabels(y_data, fontsize=9)
            ax.invert_yaxis()
            ax.set_xlabel('金额 (元)', fontsize=10)
            
            # 在柱子上显示数值
            for bar, value in zip(bars, x_data):
                width = bar.get_width()
                ax.text(width, bar.get_y() + bar.get_height()/2.,
                       f'{value:,.0f}',
                       ha='left', va='center', fontsize=8)
            
            figure.tight_layout()
            canvas.draw()
            
            self.supplier_ranking_layout.addWidget(canvas)
            
        except Exception as e:
            self.supplier_ranking_chart_placeholder.setText(f"图表加载失败: {str(e)}")
    
    def _update_type_distribution_chart(self):
        """更新发票类型分布"""
        if not self.matplotlib_available or self.df is None:
            return
        
        # 按发票类型汇总
        type_dist = self.df.groupby('发票类型').agg({
            '价税合计': 'count'
        }).reset_index()
        type_dist.columns = ['发票类型', '数量']
        
        if len(type_dist) == 0:
            self.type_distribution_chart_placeholder.setText("无数据")
            return
        
        # 创建图表
        try:
            # 清除旧的图表
            for i in reversed(range(self.type_distribution_layout.count())):
                item = self.type_distribution_layout.itemAt(i)
                if item.widget():
                    if item.widget() != self.type_distribution_label:
                        item.widget().deleteLater()
            
            # 创建新图表
            figure = self.Figure(figsize=(5, 4), dpi=100)
            canvas = self.FigureCanvas(figure)
            ax = figure.add_subplot(111)
            
            # 绘制饼图
            labels = type_dist['发票类型'].tolist()
            sizes = type_dist['数量'].tolist()
            colors = ['#27ae60', '#3498db', '#e74c3c', '#f39c12', '#9b59b6']
            
            wedges, texts, autotexts = ax.pie(sizes, labels=labels, autopct='%1.1f%%',
                                              colors=colors[:len(labels)], startangle=90)
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
                autotext.set_fontsize(8)
            
            figure.tight_layout()
            canvas.draw()
            
            self.type_distribution_layout.addWidget(canvas)
            
        except Exception as e:
            self.type_distribution_chart_placeholder.setText(f"图表加载失败: {str(e)}")
    
    def _update_charts_as_text(self):
        """用文本方式更新图表（matplotlib不可用时）"""
        # 月度趋势
        if '月份' in self.df.columns:
            monthly = self.df.groupby('月份').agg({
                '价税合计': 'sum'
            }).reset_index()
            monthly = monthly.sort_values('月份')
            trend_text = "\n".join([f"{row['月份']}: ¥{row['价税合计']:,.2f}" 
                                   for _, row in monthly.iterrows()])
            self.monthly_trend_chart_placeholder.setText(trend_text)
        
        # 销方排行榜
        supplier_stats = self.df.groupby('销方名称').agg({
            '价税合计': 'sum'
        }).reset_index()
        supplier_stats = supplier_stats.sort_values('价税合计', ascending=False).head(5)
        supplier_text = "\n".join([f"{row['销方名称']}: ¥{row['价税合计']:,.2f}" 
                                  for _, row in supplier_stats.iterrows()])
        self.supplier_ranking_chart_placeholder.setText(supplier_text)
        
        # 发票类型分布
        type_dist = self.df.groupby('发票类型').size()
        type_text = "\n".join([f"{inv_type}: {count}" 
                             for inv_type, count in type_dist.items()])
        self.type_distribution_chart_placeholder.setText(type_text)
    
    def clear_charts(self):
        """清空图表"""
        self.monthly_trend_chart_placeholder.setText("等待数据...")
        self.supplier_ranking_chart_placeholder.setText("等待数据...")
        self.type_distribution_chart_placeholder.setText("等待数据...")
    
    def refresh_data(self):
        """刷新数据"""
        parent = self.parent()
        if hasattr(parent, 'refresh_dashboard_data'):
            parent.refresh_dashboard_data()
    
    def apply_filter(self):
        """应用时间筛选"""
        # 获取年份和月份
        start_year = self.start_year_combo.currentText()
        start_month = self.start_month_combo.currentText()
        end_year = self.end_year_combo.currentText()
        end_month = self.end_month_combo.currentText()
        
        # 检查是否有选择
        if start_year == "全部" and end_year == "全部":
            # 如果没有选择筛选条件，显示所有数据
            self.update_stats(self.current_invoices)
            return
        
        try:
            # 转换为日期
            start_date = None
            end_date = None
            
            if start_year != "全部":
                start_year_int = int(start_year)
                start_month_int = int(start_month)
                start_date = datetime(start_year_int, start_month_int, 1)
            
            if end_year != "全部":
                end_year_int = int(end_year)
                end_month_int = int(end_month)
                # 结束日期为该月的最后一天
                if end_month_int == 12:
                    end_date = datetime(end_year_int + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = datetime(end_year_int, end_month_int + 1, 1) - timedelta(days=1)
            
            # 验证日期范围：结束日期不能早于开始日期
            if start_date and end_date and end_date < start_date:
                QMessageBox.warning(
                    self, 
                    "时间范围错误", 
                    f"结束时间（{end_year}年{end_month}月）不能早于开始时间（{start_year}年{start_month}月）\n\n请重新选择时间范围。"
                )
                return
            
            # 筛选发票
            filtered_invoices = []
            for invoice in self.current_invoices:
                invoice_date_str = invoice.get('开票日期', '')
                if not invoice_date_str:
                    continue
                
                # 转换发票日期
                if '年' in invoice_date_str:
                    invoice_date = datetime.strptime(invoice_date_str, '%Y年%m月%d日')
                else:
                    invoice_date = datetime.strptime(invoice_date_str, '%Y-%m-%d')
                
                # 应用筛选条件
                if start_date and invoice_date < start_date:
                    continue
                if end_date and invoice_date > end_date:
                    continue
                
                filtered_invoices.append(invoice)
            
            # 更新统计（不更新时间筛选下拉列表，避免覆盖用户选择）
            self.update_stats(filtered_invoices, update_filter_combos=False)
            
            # 显示筛选结果
            if filtered_invoices:
                self.parent().log_message(f"时间筛选完成，找到 {len(filtered_invoices)} 张发票")
            else:
                QMessageBox.information(
                    self, 
                    "筛选结果", 
                    f"在指定时间范围内没有找到发票。\n\n请尝试扩大时间范围或选择其他时间。"
                )
            
        except ValueError as e:
            QMessageBox.warning(self, "日期格式错误", f"日期格式不正确: {str(e)}")
        except Exception as e:
            QMessageBox.critical(self, "筛选失败", f"应用时间筛选时发生错误: {str(e)}")
    
    def clear_filter(self):
        """清除时间筛选"""
        if hasattr(self, 'start_year_combo'):
            self.start_year_combo.setCurrentIndex(0)  # 选择"全部"
            self.start_month_combo.setCurrentIndex(0)  # 选择"全部"
        if hasattr(self, 'end_year_combo'):
            self.end_year_combo.setCurrentIndex(0)  # 选择"全部"
            self.end_month_combo.setCurrentIndex(0)  # 选择"全部"
        # 显示所有数据
        self.update_stats(self.current_invoices)


class SettingsWidget(QWidget):
    """设置界面"""
    
    def __init__(self):
        super().__init__()
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        # 监控设置组
        monitor_group = QGroupBox("自动监控设置")
        monitor_layout = QFormLayout()
        
        self.monitor_checkbox = QCheckBox("启用自动监控")
        self.monitor_interval = QComboBox()
        self.monitor_interval.addItems(["1分钟", "5分钟", "10分钟", "30分钟", "1小时"])
        
        monitor_layout.addRow("启用监控:", self.monitor_checkbox)
        monitor_layout.addRow("监控间隔:", self.monitor_interval)
        monitor_group.setLayout(monitor_layout)
        
        # 输出设置组
        output_group = QGroupBox("输出设置")
        output_layout = QFormLayout()
        
        self.output_format = QComboBox()
        self.output_format.addItems(["Excel (.xlsx)", "CSV (.csv)", "PDF (.pdf)"])
        
        self.output_path = QLineEdit()
        self.output_path.setPlaceholderText("选择输出路径...")
        browse_btn = QPushButton("浏览")
        browse_btn.clicked.connect(self.browse_output_path)
        
        path_layout = QHBoxLayout()
        path_layout.addWidget(self.output_path)
        path_layout.addWidget(browse_btn)
        
        output_layout.addRow("输出格式:", self.output_format)
        output_layout.addRow("输出路径:", path_layout)
        output_group.setLayout(output_layout)
        
        layout.addWidget(monitor_group)
        layout.addWidget(output_group)
        layout.addStretch()
        
        self.setLayout(layout)
    
    def browse_output_path(self):
        path = QFileDialog.getExistingDirectory(self, "选择输出路径")
        if path:
            self.output_path.setText(path)


class MainWindow(QMainWindow):
    """主窗口类"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("发票归集软件 v4.2")
        self.setGeometry(100, 100, 1400, 900)
        
        # 设置窗口图标（如果有的话）
        # self.setWindowIcon(QIcon('icon.png'))
        
        self.processor = InvoiceProcessor()
        self.current_invoices = []
        self.auto_monitor_enabled = False
        self.monitor_timer = None
        
        self.init_ui()
        
    def init_ui(self):
        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 主布局
        main_layout = QVBoxLayout()
        
        # 顶部控制区域
        control_layout = QHBoxLayout()
        
        # 文件夹选择相关控件
        self.select_folder_btn = QPushButton("选择文件夹")
        self.select_folder_btn.clicked.connect(self.select_folder)
        self.select_folder_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        
        self.selected_folder_label = QLabel("未选择文件夹")
        self.selected_folder_label.setStyleSheet("color: #7f8c8d; padding: 5px;")
        
        # 扫描控制按钮
        self.scan_btn = QPushButton("开始扫描")
        self.scan_btn.clicked.connect(self.start_scan)
        self.scan_btn.setEnabled(False)
        self.scan_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
            }
        """)
        
        self.stop_btn = QPushButton("停止扫描")
        self.stop_btn.clicked.connect(self.stop_scan)
        self.stop_btn.setEnabled(False)
        self.stop_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
            }
        """)
        
        self.export_btn = QPushButton("导出Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setEnabled(False)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d35400;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
            }
        """)
        
        self.clear_btn = QPushButton("清除所有发票")
        self.clear_btn.clicked.connect(self.clear_all_invoices)
        self.clear_btn.setEnabled(False)
        self.clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
            QPushButton:disabled {
                background-color: #bdc3c7;
            }
        """)
        
        # 状态标签
        self.status_label = QLabel("就绪")
        self.status_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        
        # 添加控件到布局
        control_layout.addWidget(self.select_folder_btn)
        control_layout.addWidget(self.selected_folder_label)
        control_layout.addWidget(self.scan_btn)
        control_layout.addWidget(self.stop_btn)
        control_layout.addWidget(self.export_btn)
        control_layout.addWidget(self.clear_btn)
        control_layout.addStretch()
        control_layout.addWidget(self.status_label)
        
        # 创建选项卡
        self.tabs = QTabWidget()
        
        # 发票列表选项卡
        self.invoice_table = QTableWidget()
        self.setup_table()
        self.tabs.addTab(self.invoice_table, "发票列表")
        
        # 仪表盘选项卡
        self.dashboard = DashboardWidget()
        self.tabs.addTab(self.dashboard, "仪表盘")
        
        # 设置选项卡
        self.settings = SettingsWidget()
        self.tabs.addTab(self.settings, "设置")
        
        # 连接选项卡切换信号
        self.tabs.currentChanged.connect(self.on_tab_changed)
        
        # 日志区域
        self.log_area = QTextEdit()
        self.log_area.setMaximumHeight(150)
        self.log_area.setReadOnly(True)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        
        # 添加到主布局
        main_layout.addLayout(control_layout)
        main_layout.addWidget(self.tabs)
        main_layout.addWidget(self.progress_bar)
        main_layout.addWidget(self.log_area)
        
        central_widget.setLayout(main_layout)
        
    def setup_table(self):
        """设置表格"""
        headers = ['开票日期', '发票号码', '购方名称', '销方名称', '项目名称', '税率', '金额', '税额', '价税合计', '发票类型', '来源']
        self.invoice_table.setColumnCount(len(headers))
        self.invoice_table.setHorizontalHeaderLabels(headers)
        
        # 设置列宽策略
        header = self.invoice_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)  # 开票日期
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)  # 发票号码
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)           # 购方名称
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)           # 销方名称
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)           # 项目名称
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)  # 税率
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)  # 金额
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)  # 税额
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)  # 价税合计
        header.setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)  # 发票类型
        header.setSectionResizeMode(10, QHeaderView.ResizeMode.ResizeToContents)  # 来源
        
    def select_folder(self):
        """选择文件夹"""
        folder_path = QFileDialog.getExistingDirectory(self, "选择发票文件夹")
        if folder_path:
            self.selected_folder = folder_path
            self.selected_folder_label.setText(f"已选: {os.path.basename(folder_path)}")
            self.status_label.setText(f"已选择文件夹: {folder_path}")
            self.scan_btn.setEnabled(True)
            self.log_area.append(f"[{datetime.now().strftime('%H:%M:%S')}] 选择文件夹: {folder_path}")
            
    def start_scan(self):
        """开始扫描"""
        if hasattr(self, 'selected_folder'):
            self.scan_thread = InvoiceScannerThread(self.selected_folder)
            self.scan_thread.progress_update.connect(self.update_progress)
            self.scan_thread.status_update.connect(self.update_status)
            self.scan_thread.scan_complete.connect(self.on_scan_complete)
            self.scan_thread.log_update.connect(self.log_message)
            
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            self.scan_btn.setEnabled(False)
            self.stop_btn.setEnabled(True)
            self.status_label.setText("正在扫描...")
            
            self.scan_thread.start()
            
    def stop_scan(self):
        """停止扫描"""
        if hasattr(self, 'scan_thread'):
            # 在实际应用中，这里应该实现线程中断逻辑
            self.scan_thread.terminate()
            self.scan_thread.wait()
            
        self.progress_bar.setVisible(False)
        self.scan_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.status_label.setText("扫描已停止")
        self.log_area.append(f"[{datetime.now().strftime('%H:%M:%S')}] 扫描已停止")
    
    def update_progress(self, value):
        """更新进度"""
        self.progress_bar.setValue(value)
        
    def update_status(self, status):
        """更新状态"""
        self.status_label.setText(status)
        
    def log_message(self, message):
        """添加日志消息"""
        self.log_area.append(f"[{datetime.now().strftime('%H:%M:%S')}] {message}")
        
    def on_scan_complete(self, invoices):
        """扫描完成回调"""
        # 累加发票到现有列表（避免重复）
        existing_invoice_numbers = {inv.get('发票号码', '') for inv in self.current_invoices}
        new_invoices = [inv for inv in invoices if inv.get('发票号码', '') not in existing_invoice_numbers]
        
        # 将新发票添加到现有列表
        self.current_invoices.extend(new_invoices)
        
        # 按开票日期排序
        self.current_invoices.sort(key=lambda x: x.get('开票日期', ''), reverse=False)
        
        self.display_invoices(self.current_invoices)
        self.progress_bar.setVisible(False)
        self.scan_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.export_btn.setEnabled(True)
        self.clear_btn.setEnabled(True)  # 启用清除按钮
        
        total_count = len(self.current_invoices)
        new_count = len(new_invoices)
        self.status_label.setText(f"扫描完成，共 {total_count} 张发票（新增 {new_count} 张）")
        self.log_area.append(f"[{datetime.now().strftime('%H:%M:%S')}] 扫描完成，共 {total_count} 张发票（新增 {new_count} 张）")
        
        # 自动更新仪表盘
        self.refresh_dashboard_data()
        
    def display_invoices(self, invoices):
        """在表格中显示发票信息"""
        self.invoice_table.setRowCount(len(invoices))
        
        for row, invoice in enumerate(invoices):
            self.invoice_table.setItem(row, 0, QTableWidgetItem(invoice.get('开票日期', '')))
            self.invoice_table.setItem(row, 1, QTableWidgetItem(invoice.get('发票号码', '')))
            self.invoice_table.setItem(row, 2, QTableWidgetItem(invoice.get('购方名称', '')))
            self.invoice_table.setItem(row, 3, QTableWidgetItem(invoice.get('销方名称', '')))
            self.invoice_table.setItem(row, 4, QTableWidgetItem(invoice.get('项目名称', '')))
            self.invoice_table.setItem(row, 5, QTableWidgetItem(str(invoice.get('税率', '0.00%'))))
            self.invoice_table.setItem(row, 6, QTableWidgetItem(str(invoice.get('金额', ''))))
            self.invoice_table.setItem(row, 7, QTableWidgetItem(str(invoice.get('税额', ''))))
            self.invoice_table.setItem(row, 8, QTableWidgetItem(str(invoice.get('价税合计', ''))))
            self.invoice_table.setItem(row, 9, QTableWidgetItem(invoice.get('发票类型', '')))
            self.invoice_table.setItem(row, 10, QTableWidgetItem(invoice.get('文件路径', '')))
    
    def _analyze_data(self, ws):
        """分析Sheet1数据结构"""
        # 获取数据范围
        max_row = ws.max_row
        max_col = ws.max_column
        
        # 获取列标题
        headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]
        
        # 查找关键列索引
        col_map = {}
        for i, h in enumerate(headers, 1):
            if h and isinstance(h, str):
                col_map[h.strip()] = i
        
        return {
            'max_row': max_row,
            'max_col': max_col,
            'headers': headers,
            'col_map': col_map
        }
    
    def _get_unique_values(self, ws, col_idx, start_row, end_row):
        """获取某列的唯一值列表"""
        values = {}
        for r in range(start_row, end_row + 1):
            v = ws.cell(r, col_idx).value
            if v:
                values[str(v).strip()] = values.get(str(v).strip(), 0) + 1
        return list(values.keys())
    
    def _get_top_sellers(self, ws, col_idx, amt_col_idx, start_row, end_row, top_n=10):
        """获取销方TOP N"""
        seller_amts = {}
        for r in range(start_row, end_row + 1):
            seller = ws.cell(r, col_idx).value
            amt = ws.cell(r, amt_col_idx).value
            if seller and amt:
                seller = str(seller).strip()
                try:
                    amt = float(amt)
                except:
                    amt = 0
                seller_amts[seller] = seller_amts.get(seller, 0) + amt
        
        # 排序取TOP N
        sorted_sellers = sorted(seller_amts.items(), key=lambda x: x[1], reverse=True)[:top_n]
        return [s[0] for s in sorted_sellers]
    
    def _convert_numeric_columns(self, wb, data_info):
        """确保金额/税额/价税合计为数值类型"""
        ws = wb['Sheet1']
        max_row = data_info['max_row']
        col_map = data_info['col_map']
        
        for col_name in ['金额', '税额', '价税合计']:
            col_idx = col_map.get(col_name)
            if col_idx:
                for r in range(2, max_row + 1):
                    val = ws.cell(r, col_idx).value
                    if isinstance(val, str):
                        try:
                            ws.cell(r, col_idx).value = float(val.strip())
                        except:
                            pass
    
    def _create_dashboard(self, wb, data_info):
        """创建动态仪表盘"""
        ws = wb.create_sheet('动态仪表盘')
        
        max_row = data_info['max_row']
        col_map = data_info['col_map']
        ws1 = wb['Sheet1']
        
        # 获取列索引（现在包含年份和月份字段）
        date_col = col_map.get('开票日期', 1)
        invoice_col = col_map.get('发票号码', 2)
        buyer_col = col_map.get('购方名称', 3)
        seller_col = col_map.get('销方名称', 4)
        item_col = col_map.get('项目名称', 5)
        rate_col = col_map.get('税率', 6)
        amt_col = col_map.get('金额', 7)
        tax_col = col_map.get('税额', 8)
        total_col = col_map.get('价税合计', 9)
        type_col = col_map.get('发票类型', 10)
        year_col = col_map.get('年份', 11)
        month_col = col_map.get('月份', 12)
        
        # 样式定义
        title_font = ExcelFont(name='微软雅黑', size=16, bold=True, color='FFFFFF')
        header_font = ExcelFont(name='微软雅黑', size=11, bold=True)
        kpi_font = ExcelFont(name='微软雅黑', size=14, bold=True, color='2F5496')
        normal_font = ExcelFont(name='微软雅黑', size=10)
        
        title_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_fill = PatternFill(start_color='D6DCE5', end_color='D6DCE5', fill_type='solid')
        kpi_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # ========== 1. 标题 ==========
        ws.merge_cells('A1:L1')
        ws['A1'] = '📊 发票数据动态仪表盘'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws['A1'].alignment = center_align
        ws.row_dimensions[1].height = 35
        
        # ========== 2. KPI 指标卡 ==========
        kpi_items = [
            ('B', '💰 总价税合计', f'=SUM(Sheet1!${chr(64+total_col)}$2:${chr(64+total_col)}${max_row})'),
            ('D', '💵 发票总金额', f'=SUM(Sheet1!${chr(64+amt_col)}$2:${chr(64+amt_col)}${max_row})'),
            ('F', '🧾 税额合计', f'=SUM(Sheet1!${chr(64+tax_col)}$2:${chr(64+tax_col)}${max_row})'),
            ('H', '📋 发票张数', f'=COUNTA(Sheet1!${chr(64+invoice_col)}$2:${chr(64+invoice_col)}${max_row})'),
            ('J', '🏢 供应商数', f'=SUMPRODUCT(1/COUNTIF(Sheet1!${chr(64+seller_col)}$2:${chr(64+seller_col)}${max_row},Sheet1!${chr(64+seller_col)}$2:${chr(64+seller_col)}${max_row}))'),
            ('L', '📦 项目品类', f'=SUMPRODUCT(1/COUNTIF(Sheet1!${chr(64+item_col)}$2:${chr(64+item_col)}${max_row},Sheet1!${chr(64+item_col)}$2:${chr(64+item_col)}${max_row}))'),
        ]
        
        for col, label, formula in kpi_items:
            ws[f'{col}3'] = label
            ws[f'{col}3'].font = header_font
            ws[f'{col}3'].fill = header_fill
            ws[f'{col}3'].alignment = center_align
            ws[f'{col}3'].border = thin_border
            
            ws[f'{col}5'] = formula
            ws[f'{col}5'].font = ExcelFont(name='微软雅黑', size=10, bold=True, color='2F5496')
            ws[f'{col}5'].fill = kpi_fill
            ws[f'{col}5'].alignment = center_align
            ws[f'{col}5'].border = thin_border
            if col in ['B', 'D', 'F']:
                ws[f'{col}5'].number_format = '#,##0'
            else:
                ws[f'{col}5'].number_format = '0'
        
        ws.row_dimensions[3].height = 22
        ws.row_dimensions[5].height = 28
        
        # ========== 3. 月度趋势分析 ==========
        ws['B7'] = '📈 月度趋势分析'
        ws['B7'].font = ExcelFont(name='微软雅黑', size=12, bold=True, color='2F5496')
        
        month_headers = ['月份', '发票张数', '金额', '税额', '价税合计', '环比增幅']
        for i, h in enumerate(month_headers):
            cell = ws.cell(row=9, column=2+i)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # 获取所有年月
        year_month_data = {}
        for r in range(2, max_row + 1):
            date_val = ws1.cell(r, date_col).value
            if date_val:
                try:
                    if isinstance(date_val, str):
                        dt = datetime.strptime(date_val.split()[0], '%Y-%m-%d')
                    else:
                        dt = date_val
                    ym = (dt.year, dt.month)
                    year_month_data[ym] = True
                except:
                    pass
        
        sorted_ym = sorted(year_month_data.keys())
        
        # 写入月度数据
        row_idx = 10
        for year, month in sorted_ym:
            ws.cell(row=row_idx, column=2).value = f'{year}-{month:02d}'
            ws.cell(row=row_idx, column=2).alignment = center_align
            ws.cell(row=row_idx, column=2).border = thin_border
            
            ws.cell(row=row_idx, column=3).value = f'=SUMPRODUCT((Sheet1!${chr(64+year_col)}$2:${chr(64+year_col)}${max_row}={year})*(Sheet1!${chr(64+month_col)}$2:${chr(64+month_col)}${max_row}={month}))'
            ws.cell(row=row_idx, column=3).alignment = center_align
            ws.cell(row=row_idx, column=3).border = thin_border
            
            ws.cell(row=row_idx, column=4).value = f'=SUMIFS(Sheet1!${chr(64+amt_col)}$2:${chr(64+amt_col)}${max_row},Sheet1!${chr(64+year_col)}$2:${chr(64+year_col)}${max_row},{year},Sheet1!${chr(64+month_col)}$2:${chr(64+month_col)}${max_row},{month})'
            ws.cell(row=row_idx, column=4).number_format = '#,##0'
            ws.cell(row=row_idx, column=4).alignment = center_align
            ws.cell(row=row_idx, column=4).border = thin_border
            
            ws.cell(row=row_idx, column=5).value = f'=SUMIFS(Sheet1!${chr(64+tax_col)}$2:${chr(64+tax_col)}${max_row},Sheet1!${chr(64+year_col)}$2:${chr(64+year_col)}${max_row},{year},Sheet1!${chr(64+month_col)}$2:${chr(64+month_col)}${max_row},{month})'
            ws.cell(row=row_idx, column=5).number_format = '#,##0'
            ws.cell(row=row_idx, column=5).alignment = center_align
            ws.cell(row=row_idx, column=5).border = thin_border
            
            ws.cell(row=row_idx, column=6).value = f'=SUMIFS(Sheet1!${chr(64+total_col)}$2:${chr(64+total_col)}${max_row},Sheet1!${chr(64+year_col)}$2:${chr(64+year_col)}${max_row},{year},Sheet1!${chr(64+month_col)}$2:${chr(64+month_col)}${max_row},{month})'
            ws.cell(row=row_idx, column=6).number_format = '#,##0'
            ws.cell(row=row_idx, column=6).alignment = center_align
            ws.cell(row=row_idx, column=6).border = thin_border
            
            if row_idx > 10:
                ws.cell(row=row_idx, column=7).value = f'=IF(F{row_idx-1}=0,"-",F{row_idx}/F{row_idx-1}-1)'
            else:
                ws.cell(row=row_idx, column=7).value = '-'
            ws.cell(row=row_idx, column=7).number_format = '0.0%'
            ws.cell(row=row_idx, column=7).alignment = center_align
            ws.cell(row=row_idx, column=7).border = thin_border
            
            row_idx += 1
        
        # 合计行
        total_row = row_idx
        ws.cell(row=total_row, column=2).value = '合计'
        ws.cell(row=total_row, column=2).font = header_font
        ws.cell(row=total_row, column=2).fill = header_fill
        ws.cell(row=total_row, column=2).alignment = center_align
        ws.cell(row=total_row, column=2).border = thin_border
        
        for col in range(3, 7):
            ws.cell(row=total_row, column=col).value = f'=SUM({chr(64+col)}10:{chr(64+col)}{total_row-1})'
            ws.cell(row=total_row, column=col).font = header_font
            ws.cell(row=total_row, column=col).fill = header_fill
            ws.cell(row=total_row, column=col).alignment = center_align
            ws.cell(row=total_row, column=col).border = thin_border
            if col >= 4:
                ws.cell(row=total_row, column=col).number_format = '#,##0'
        
        ws.cell(row=total_row, column=7).value = '-'
        ws.cell(row=total_row, column=7).font = header_font
        ws.cell(row=total_row, column=7).fill = header_fill
        ws.cell(row=total_row, column=7).alignment = center_align
        ws.cell(row=total_row, column=7).border = thin_border
        
        # ========== 4. 购方分析 ==========
        section_start = total_row + 3
        
        ws[f'B{section_start}'] = '🏢 购方分析'
        ws[f'B{section_start}'].font = ExcelFont(name='微软雅黑', size=12, bold=True, color='2F5496')
        
        buyers = self._get_unique_values(ws1, buyer_col, 2, max_row)
        
        for i, h in enumerate(['购方名称', '金额', '占比']):
            cell = ws.cell(row=section_start+1, column=2+i)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        for idx, buyer in enumerate(buyers):
            r = section_start + 2 + idx
            ws.cell(row=r, column=2).value = buyer
            ws.cell(row=r, column=2).alignment = left_align
            ws.cell(row=r, column=2).border = thin_border
            
            ws.cell(row=r, column=3).value = f'=SUMIF(Sheet1!${chr(64+buyer_col)}$2:${chr(64+buyer_col)}${max_row},B{r},Sheet1!${chr(64+total_col)}$2:${chr(64+total_col)}${max_row})'
            ws.cell(row=r, column=3).number_format = '#,##0'
            ws.cell(row=r, column=3).alignment = center_align
            ws.cell(row=r, column=3).border = thin_border
            
            ws.cell(row=r, column=4).value = f'=C{r}/$B$5'
            ws.cell(row=r, column=4).number_format = '0.0%'
            ws.cell(row=r, column=4).alignment = center_align
            ws.cell(row=r, column=4).border = thin_border
        
        # ========== 5. 销方TOP10 ==========
        ws[f'G{section_start}'] = '🛒 销方 TOP10'
        ws[f'G{section_start}'].font = ExcelFont(name='微软雅黑', size=12, bold=True, color='2F5496')
        
        sellers = self._get_top_sellers(ws1, seller_col, total_col, 2, max_row, 10)
        
        for i, h in enumerate(['销方名称', '金额', '占比']):
            cell = ws.cell(row=section_start+1, column=7+i)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        for idx, seller in enumerate(sellers):
            r = section_start + 2 + idx
            ws.cell(row=r, column=7).value = seller
            ws.cell(row=r, column=7).alignment = left_align
            ws.cell(row=r, column=7).border = thin_border
            
            ws.cell(row=r, column=8).value = f'=SUMIF(Sheet1!${chr(64+seller_col)}$2:${chr(64+seller_col)}${max_row},G{r},Sheet1!${chr(64+total_col)}$2:${chr(64+total_col)}${max_row})'
            ws.cell(row=r, column=8).number_format = '#,##0'
            ws.cell(row=r, column=8).alignment = center_align
            ws.cell(row=r, column=8).border = thin_border
            
            ws.cell(row=r, column=9).value = f'=H{r}/$B$5'
            ws.cell(row=r, column=9).number_format = '0.0%'
            ws.cell(row=r, column=9).alignment = center_align
            ws.cell(row=r, column=9).border = thin_border
        
        # ========== 6. 税率分布 ==========
        max_partner_rows = max(len(buyers), len(sellers))
        rate_section = section_start + max_partner_rows + 4
        
        ws[f'B{rate_section}'] = '📊 税率分布'
        ws[f'B{rate_section}'].font = ExcelFont(name='微软雅黑', size=12, bold=True, color='2F5496')
        
        rates = self._get_unique_values(ws1, rate_col, 2, max_row)
        
        for i, h in enumerate(['税率', '金额', '占比']):
            cell = ws.cell(row=rate_section+1, column=2+i)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        for idx, rate in enumerate(rates):
            r = rate_section + 2 + idx
            ws.cell(row=r, column=2).value = rate
            ws.cell(row=r, column=2).alignment = center_align
            ws.cell(row=r, column=2).border = thin_border
            
            ws.cell(row=r, column=3).value = f'=SUMIF(Sheet1!${chr(64+rate_col)}$2:${chr(64+rate_col)}${max_row},B{r},Sheet1!${chr(64+total_col)}$2:${chr(64+total_col)}${max_row})'
            ws.cell(row=r, column=3).number_format = '#,##0'
            ws.cell(row=r, column=3).alignment = center_align
            ws.cell(row=r, column=3).border = thin_border
            
            ws.cell(row=r, column=4).value = f'=C{r}/$B$5'
            ws.cell(row=r, column=4).number_format = '0.0%'
            ws.cell(row=r, column=4).alignment = center_align
            ws.cell(row=r, column=4).border = thin_border
        
        # ========== 7. 发票类型 ==========
        ws[f'G{rate_section}'] = '📋 发票类型'
        ws[f'G{rate_section}'].font = ExcelFont(name='微软雅黑', size=12, bold=True, color='2F5496')
        
        types = self._get_unique_values(ws1, type_col, 2, max_row)
        
        for i, h in enumerate(['发票类型', '金额', '占比']):
            cell = ws.cell(row=rate_section+1, column=7+i)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        for idx, inv_type in enumerate(types):
            r = rate_section + 2 + idx
            ws.cell(row=r, column=7).value = inv_type
            ws.cell(row=r, column=7).alignment = center_align
            ws.cell(row=r, column=7).border = thin_border
            
            ws.cell(row=r, column=8).value = f'=SUMIF(Sheet1!${chr(64+type_col)}$2:${chr(64+type_col)}${max_row},G{r},Sheet1!${chr(64+total_col)}$2:${chr(64+total_col)}${max_row})'
            ws.cell(row=r, column=8).number_format = '#,##0'
            ws.cell(row=r, column=8).alignment = center_align
            ws.cell(row=r, column=8).border = thin_border
            
            ws.cell(row=r, column=9).value = f'=H{r}/$B$5'
            ws.cell(row=r, column=9).number_format = '0.0%'
            ws.cell(row=r, column=9).alignment = center_align
            ws.cell(row=r, column=9).border = thin_border
        
        # ========== 设置列宽和字体 ==========
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 32
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 14
        ws.column_dimensions['K'].width = 3
        ws.column_dimensions['L'].width = 14
        
        for row in range(1, 50):
            ws.row_dimensions[row].height = 18
    
    def export_to_excel(self):
        """导出到Excel - 优化版，支持固定列宽，自动生成仪表盘，包含年份和月份字段"""
        if self.current_invoices:
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "保存Excel文件", 
                f"发票清单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel 文件 (*.xlsx)"
            )
            
            if file_path:
                try:
                    # 定义字段顺序，添加年份和月份字段
                    column_order = ['开票日期', '发票号码', '购方名称', '销方名称', '项目名称', '税率', '金额', '税额', '价税合计', '发票类型', '年份', '月份', '来源']
                    
                    # 确保每张发票都有所有字段，缺失的字段使用默认值
                    for invoice in self.current_invoices:
                        for col in column_order:
                            if col not in invoice or not invoice[col]:
                                if col in ['金额', '税额', '价税合计']:
                                    invoice[col] = '0.00'
                                elif col == '税率':
                                    invoice[col] = '0.00%'
                                elif col in ['年份', '月份']:
                                    invoice[col] = 0
                                elif col == '来源':
                                    invoice[col] = invoice.get('文件路径', '')
                                else:
                                    invoice[col] = '未识别'
                        
                        # 从开票日期提取年份和月份
                        date_str = invoice.get('开票日期', '')
                        if date_str and date_str != '未识别':
                            try:
                                if '年' in date_str:
                                    dt = datetime.strptime(date_str, '%Y年%m月%d日')
                                else:
                                    dt = datetime.strptime(date_str, '%Y-%m-%d')
                                invoice['年份'] = dt.year
                                invoice['月份'] = dt.month
                            except:
                                invoice['年份'] = 0
                                invoice['月份'] = 0
                    
                    # 创建DataFrame并按指定顺序排列字段
                    df = pd.DataFrame(self.current_invoices, columns=column_order)
                    
                    # 按开票日期升序排列
                    df = df.sort_values(by=['开票日期'])
                    
                    # 使用openpyxl导出并设置格式
                    from openpyxl import load_workbook
                    from openpyxl.utils import get_column_letter
                    
                    # 先使用pandas导出
                    df.to_excel(file_path, index=False, sheet_name='Sheet1')
                    
                    # 加载工作簿并设置格式
                    wb = load_workbook(file_path)
                    ws = wb.active
                    
                    # 定义列宽（包含年份和月份字段）
                    column_widths = {
                        'A': 15,  # 开票日期
                        'B': 25,  # 发票号码
                        'C': 30,  # 购方名称
                        'D': 30,  # 销方名称
                        'E': 50,  # 项目名称
                        'F': 10,  # 税率
                        'G': 12,  # 金额
                        'H': 12,  # 税额
                        'I': 15,  # 价税合计
                        'J': 15,  # 发票类型
                        'K': 10,  # 年份
                        'L': 10,  # 月份
                        'M': 60   # 来源
                    }
                    
                    # 设置列宽
                    for col, width in column_widths.items():
                        ws.column_dimensions[col].width = width
                    
                    # 设置表头样式
                    header_font = ExcelFont(name='微软雅黑', size=11, bold=True, color='FFFFFF')
                    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # 应用表头样式
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    
                    # 设置数据行样式
                    data_font = ExcelFont(name='微软雅黑', size=10)
                    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # 金额、税额、价税合计、年份、月份列居中对齐
                    for row in ws.iter_rows(min_row=2):
                        for idx, cell in enumerate(row, 1):
                            cell.font = data_font
                            cell.border = thin_border
                            
                            # 第7、8、9列（税率、金额、税额）和第10列（价税合计）右对齐
                            # 第11、12列（年份、月份）居中对齐
                            # 第13列（来源）左对齐
                            if idx in [6, 7, 8, 9]:
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                            elif idx in [10, 11]:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            elif idx == 13:
                                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            else:
                                cell.alignment = data_alignment
                    
                    # 冻结首行
                    ws.freeze_panes = 'A2'
                    
                    # 生成智能仪表盘
                    self.log_message("正在生成智能仪表盘...")
                    
                    # 分析数据
                    data_info = self._analyze_data(ws)
                    
                    # 转换数值列
                    self._convert_numeric_columns(wb, data_info)
                    
                    # 创建仪表盘（不再添加辅助列，因为年份和月份已经在数据中）
                    self._create_dashboard(wb, data_info)
                    
                    # 保存工作簿
                    wb.save(file_path)
                    
                    self.log_message(f"成功导出到: {file_path}")
                    QMessageBox.information(
                        self, 
                        "导出成功", 
                        f"发票清单已成功导出到:\n{file_path}\n\n共 {len(self.current_invoices)} 张发票\n\n已自动生成智能仪表盘（动态仪表盘工作表）"
                    )
                except Exception as e:
                    import traceback
                    error_msg = f"导出失败: {str(e)}\n\n详细错误:\n{traceback.format_exc()}"
                    self.log_message(error_msg)
                    QMessageBox.critical(
                        self, 
                        "导出失败", 
                        error_msg
                    )
    
    def refresh_dashboard_data(self):
        """刷新仪表盘数据"""
        if self.current_invoices:
            # 首次加载时更新时间筛选下拉列表
            self.dashboard.update_stats(self.current_invoices, update_filter_combos=True)
            self.log_message("仪表盘数据已刷新")
    
    def clear_all_invoices(self):
        """清除所有发票"""
        if not self.current_invoices:
            QMessageBox.information(self, "提示", "当前没有发票数据")
            return
        
        # 确认对话框
        reply = QMessageBox.question(
            self, 
            "确认清除", 
            f"确定要清除所有 {len(self.current_invoices)} 张发票数据吗？\n此操作不可恢复！",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # 清空发票列表
            self.current_invoices = []
            
            # 清空表格
            self.invoice_table.setRowCount(0)
            
            # 清空仪表盘
            self.dashboard.update_stats([])
            
            # 更新按钮状态
            self.export_btn.setEnabled(False)
            self.clear_btn.setEnabled(False)
            
            # 更新状态
            self.status_label.setText("已清除所有发票")
            self.log_message(f"[{datetime.now().strftime('%H:%M:%S')}] 已清除所有发票数据")
    
    def on_tab_changed(self, index):
        """选项卡切换事件"""
        # 当切换到仪表盘选项卡时，自动刷新数据
        if index == 1:  # 仪表盘选项卡
            self.refresh_dashboard_data()